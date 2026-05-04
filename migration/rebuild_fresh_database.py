import argparse
import re
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

MIGRATED_BY = "migration@mycitibidadi.com"
NOW = datetime.now().isoformat(timespec="seconds")
TODAY = date.today().isoformat()

SITE_TYPE_MAP = {
    "30x40": 1200,
    "40x60": 2400,
    "30x50": 1500,
    "50x80": 4000,
    "27x50": 1350,
    "40x40": 1600,
    "60x80": 4800,
    "30x60": 1800,
    "40x50": 2000,
}

SITES_COLS = [
    "SiteID", "SiteNo", "Phase", "Released", "SiteType", "Sizesqft", "RegDate",
    "AttachmentURLs", "FlaggedForAttention", "FlagComment", "FlaggedBy", "FlaggedAt",
    "IsDeleted", "CreatedBy", "CreatedAt", "ModifiedBy", "ModifiedAt",
]

PEOPLE_COLS = [
    "PersonID", "FullName", "Mobile1", "Mobile2", "Email", "Address", "CommAddress",
    "PhotoURL", "IDProofURLs", "Notes", "IsDeleted", "CreatedBy", "CreatedAt",
    "ModifiedBy", "ModifiedAt",
]

OWNERS_COLS = [
    "OwnerID", "SiteID", "PersonID", "MembershipNo", "MemberSince", "IsCurrent",
    "OwnershipStartDate", "OwnershipEndDate", "NominatedContact", "IsCouncilMember",
    "AgentID", "Status", "FlaggedForAttention", "FlagComment", "FlaggedBy", "FlaggedAt",
    "Notes", "IsDeleted", "CreatedBy", "CreatedAt", "ModifiedBy", "ModifiedAt",
]

AGENTS_COLS = [
    "AgentID", "Name", "Mobile", "Email", "PhotoURL", "IDProofURLs", "Notes",
    "IsDeleted", "CreatedBy", "CreatedAt", "ModifiedBy", "ModifiedAt",
]

PAYMENTS_COLS = [
    "PaymentID", "SiteID", "OwnerID", "HeadID", "Amount", "Mode", "PaymentDate",
    "ReceiptNo", "BankRef", "ProofURL", "FlaggedForAttention", "FlagComment", "FlaggedBy", "FlaggedAt",
    "RecordedBy", "RecordedAt", "IsDeleted",
    "CreatedBy", "CreatedAt", "ModifiedBy", "ModifiedAt",
]

TRANSFERS_COLS = [
    "TransferID", "SiteID", "FromOwnerID", "ToOwnerID", "TransferDate", "SalePrice",
    "DocRef", "RecordedBy", "RecordedAt", "IsDeleted", "CreatedBy", "CreatedAt",
    "ModifiedBy", "ModifiedAt",
]

CALLLOG_COLS = [
    "LogID", "SiteID", "OwnerID", "LogDate", "CalledBy", "Summary", "FollowUpAction",
    "AssignedTo", "AssignedToName", "FollowUpDone", "DoneBy", "DoneAt", "LoggedAt",
    "IsDeleted", "CreatedBy", "CreatedAt", "ModifiedBy", "ModifiedAt",
]

AUDIT_COLS = [
    "AuditID", "Timestamp", "UserEmail", "DisplayName", "Action", "Tab", "RecordID",
    "FieldName", "OldValue", "NewValue",
]


def clean_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def normalize_name(v):
    s = clean_str(v)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def name_tokens(v):
    s = normalize_name(v).lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    toks = [t for t in s.split() if t and t not in {"mr", "mrs", "ms", "dr", "w", "o", "s", "d"}]
    return set(toks)


def names_match(a, b):
    ta = name_tokens(a)
    tb = name_tokens(b)
    if not ta or not tb:
        return False
    if ta == tb:
        return True
    overlap = len(ta & tb)
    min_len = min(len(ta), len(tb))
    return overlap >= max(1, min_len - 1)


def normalize_site_no(v):
    s = clean_str(v).upper().replace(" ", "")
    if not s:
        return ""
    m = re.match(r"^(\d+)([A-Z]+)$", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return s


def parse_phase(v):
    s = clean_str(v)
    if not s:
        return ""
    m = re.search(r"([12])", s)
    return int(m.group(1)) if m else ""


def parse_site_phase(value):
    s = clean_str(value)
    m = re.match(r"^\s*([0-9]+(?:-[A-Za-z]+|[A-Za-z]+)?)\s*/\s*([12])\s*$", s)
    if not m:
        return "", ""
    return normalize_site_no(m.group(1)), int(m.group(2))


def parse_site_phase_list(value):
    s = clean_str(value)
    if not s:
        return []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if not parts:
        parts = [s]
    pairs = []
    for p in parts:
        site_no, phase = parse_site_phase(p)
        if site_no and phase in (1, 2):
            pairs.append((site_no, phase))
    return pairs


def clean_mobile(v):
    s = clean_str(v)
    if not s:
        return ""
    s = re.sub(r"[^0-9+]", "", s)
    s = re.sub(r"^(\+91|91)", "", s)
    if len(s) > 10:
        s = s[-10:]
    return s


def split_mobiles(v):
    s = clean_str(v)
    if not s:
        return "", ""
    nums = re.findall(r"(?:\+91)?\d[\d\s\-]{8,}\d", s)
    cleaned = []
    for n in nums:
        c = clean_mobile(n)
        if c and c not in cleaned:
            cleaned.append(c)
    if not cleaned:
        c = clean_mobile(s)
        if c:
            cleaned.append(c)
    return (cleaned[0] if cleaned else "", cleaned[1] if len(cleaned) > 1 else "")


def to_date(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = clean_str(v)
    if not s:
        return ""
    for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s


def parse_plot_size(plot_size, area):
    p = clean_str(plot_size).lower().replace(" ", "")
    if p in SITE_TYPE_MAP:
        expected = SITE_TYPE_MAP[p]
        try:
            area_num = int(float(area))
            if area_num > 0:
                return p, area_num
        except Exception:
            return p, expected
    try:
        area_num = int(float(area))
        if area_num > 0:
            return "non-standard", area_num
    except Exception:
        pass
    return "non-standard", ""


def split_names(name_text):
    raw = normalize_name(name_text)
    if not raw:
        return []
    if raw.upper() == "UNREGISTERED":
        return []
    parts = [p.strip() for p in re.split(r"\s*(?:&|,| and )\s*", raw, flags=re.IGNORECASE) if p.strip()]
    return parts if parts else [raw]


def next_id(prefix, n):
    return f"{prefix}{str(n).zfill(4)}"


def to_float(v):
    s = clean_str(v)
    if not s:
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def build_database(base_db_path, bank_csv_path, owners_list_path, members_list_path, output_path):
    base_db_path = Path(base_db_path)
    bank_csv_path = Path(bank_csv_path)
    owners_list_path = Path(owners_list_path)
    members_list_path = Path(members_list_path)
    output_path = Path(output_path)

    payment_heads_df = pd.read_excel(base_db_path, sheet_name="PaymentHeads")
    roles_df = pd.read_excel(base_db_path, sheet_name="Roles")

    # Load old DB site reference — used for site-no resolution and attribute cross-check.
    old_sites_ref = pd.read_excel(base_db_path, sheet_name="Sites")
    old_site_keys = set()
    old_site_data = {}
    for _, _orow in old_sites_ref.iterrows():
        _sno = clean_str(_orow.get("SiteNo", ""))
        try:
            _ph = int(_orow.get("Phase", 0))
        except Exception:
            continue
        if _sno and _ph in (1, 2):
            old_site_keys.add((_sno, _ph))
            old_site_data[(_sno, _ph)] = {
                "SiteType": clean_str(_orow.get("SiteType", "")),
                "Sizesqft": _orow.get("Sizesqft", ""),
            }

    # Keep existing heads and append an Additional head if missing.
    if "HeadName" in payment_heads_df.columns:
        has_additional = payment_heads_df["HeadName"].astype(str).str.strip().str.lower().eq("additional").any()
    else:
        has_additional = False

    additional_head_id = "H005"
    if not has_additional:
        existing_ids = payment_heads_df.get("HeadID", pd.Series([], dtype=str)).astype(str)
        nums = []
        for hid in existing_ids:
            m = re.match(r"^H(\d+)$", hid.strip())
            if m:
                nums.append(int(m.group(1)))
        next_num = (max(nums) + 1) if nums else 5
        additional_head_id = f"H{str(next_num).zfill(3)}"
        add_row = {
            "HeadID": additional_head_id,
            "HeadName": "Additional",
            "AmountType": "Flat",
            "ExpectedAmountFlat": "",
            "ExpectedAmountPerSqft": "",
            "DueDate": "",
            "IsActive": "TRUE",
            "Notes": "Additional / adjustment receipts",
            "IsDeleted": "FALSE",
            "CreatedBy": MIGRATED_BY,
            "CreatedAt": NOW,
            "ModifiedBy": "",
            "ModifiedAt": "",
        }
        # Add only keys that exist in current schema.
        add_row = {k: add_row.get(k, "") for k in payment_heads_df.columns}
        payment_heads_df = pd.concat([payment_heads_df, pd.DataFrame([add_row])], ignore_index=True)
    else:
        additional_row = payment_heads_df[payment_heads_df["HeadName"].astype(str).str.strip().str.lower() == "additional"].iloc[0]
        additional_head_id = str(additional_row.get("HeadID", "H005"))

    members_df = pd.read_excel(members_list_path, sheet_name="Final List")
    bank_df = pd.read_csv(bank_csv_path)

    wb = load_workbook(owners_list_path, data_only=True)
    ws = wb["Owner List Ph1&Ph2"]

    headers = [clean_str(ws.cell(2, c).value) for c in range(1, 13)]
    hidx = {h: i + 1 for i, h in enumerate(headers)}

    sites = []
    site_key_to_id = {}
    site_id_to_row = {}

    people = []
    person_key_to_id = {}

    owners = []
    owners_by_site = {}

    warnings = []

    def resolve_site_no(site_no, phase):
        """If site_no has an A/B suffix but old DB had only the bare base number, use the base.

        Example: source has '374-A' and '374-B' but old DB had just '374' → both resolve to '374'
        so they become co-owners of one site instead of two separate sites.
        Leaves site_no unchanged when old DB already tracked the A/B variant as its own site.
        """
        m = re.match(r"^(\d+)-([A-Z]+)$", site_no)
        if m:
            base = m.group(1)
            if (base, phase) in old_site_keys and (site_no, phase) not in old_site_keys:
                return base
        return site_no

    def get_or_create_site(site_no, phase, plot_size, area, reg_date):
        key = f"{site_no}-P{phase}"
        if key in site_key_to_id:
            return site_key_to_id[key]
        site_id = next_id("S", len(sites) + 1)
        stype, sqft = parse_plot_size(plot_size, area)
        row = {
            "SiteID": site_id,
            "SiteNo": site_no,
            "Phase": phase,
            "Released": "TRUE",
            "SiteType": stype,
            "Sizesqft": sqft,
            "RegDate": to_date(reg_date),
            "AttachmentURLs": "",
            "FlaggedForAttention": "FALSE",
            "FlagComment": "",
            "FlaggedBy": "",
            "FlaggedAt": "",
            "IsDeleted": "FALSE",
            "CreatedBy": MIGRATED_BY,
            "CreatedAt": NOW,
            "ModifiedBy": "",
            "ModifiedAt": "",
        }
        sites.append(row)
        site_key_to_id[key] = site_id
        site_id_to_row[site_id] = row
        return site_id

    def get_or_create_person(full_name, mobile1, mobile2, email):
        full_name = normalize_name(full_name)
        mobile1 = clean_mobile(mobile1)
        mobile2 = clean_mobile(mobile2)
        email = clean_str(email)
        key = (full_name.lower(), mobile1, mobile2, email.lower())
        if key in person_key_to_id:
            return person_key_to_id[key]
        person_id = next_id("P", len(people) + 1)
        people.append({
            "PersonID": person_id,
            "FullName": full_name,
            "Mobile1": mobile1,
            "Mobile2": mobile2,
            "Email": email,
            "Address": "",
            "CommAddress": "",
            "PhotoURL": "",
            "IDProofURLs": "",
            "Notes": "",
            "IsDeleted": "FALSE",
            "CreatedBy": MIGRATED_BY,
            "CreatedAt": NOW,
            "ModifiedBy": "",
            "ModifiedAt": "",
        })
        person_key_to_id[key] = person_id
        return person_id

    def add_owner(site_id, person_id, membership_no, member_since, is_current):
        owner_id = next_id("O", len(owners) + 1)
        row = {
            "OwnerID": owner_id,
            "SiteID": site_id,
            "PersonID": person_id,
            "MembershipNo": clean_str(membership_no),
            "MemberSince": to_date(member_since),
            "IsCurrent": "TRUE" if is_current else "FALSE",
            "OwnershipStartDate": "",
            "OwnershipEndDate": "" if is_current else TODAY,
            "NominatedContact": "",
            "IsCouncilMember": "FALSE",
            "AgentID": "",
            "Status": "Active" if clean_str(membership_no) else "Non-member",
            "FlaggedForAttention": "FALSE",
            "FlagComment": "",
            "FlaggedBy": "",
            "FlaggedAt": "",
            "Notes": "",
            "IsDeleted": "FALSE",
            "CreatedBy": MIGRATED_BY,
            "CreatedAt": NOW,
            "ModifiedBy": "",
            "ModifiedAt": "",
        }
        owners.append(row)
        owners_by_site.setdefault(site_id, []).append(row)
        return row

    # Build Sites, People, Owners from Owners List (authoritative source).
    merged_from = {}  # site_id -> set of original source site_nos that were consolidated into it
    for r in range(3, ws.max_row + 1):
        site_no = normalize_site_no(ws.cell(r, hidx["Site No"]).value)
        phase = parse_phase(ws.cell(r, hidx["Phase"]).value)

        if not site_no or phase not in (1, 2):
            site_ph = clean_str(ws.cell(r, hidx["Site/Ph"]).value)
            if site_ph:
                site_no, phase = parse_site_phase(site_ph)

        if not site_no or phase not in (1, 2):
            continue

        site_no_original = site_no
        site_no = resolve_site_no(site_no, phase)
        site_id = get_or_create_site(
            site_no=site_no,
            phase=phase,
            plot_size=ws.cell(r, hidx["Plot Size"]).value,
            area=ws.cell(r, hidx["Site Area(sft)"]).value,
            reg_date=ws.cell(r, hidx["Site Reg Dt"]).value,
        )
        if site_no != site_no_original:
            merged_from.setdefault(site_id, set()).add(site_no_original)

        name_cell = ws.cell(r, hidx["Name"])
        name_raw = clean_str(name_cell.value)
        if not name_raw or name_raw.upper() == "UNREGISTERED":
            continue

        is_struck = bool(name_cell.font and name_cell.font.strike)
        contact_raw = ws.cell(r, hidx["Contact no. given"]).value
        mobile1, mobile2 = split_mobiles(contact_raw)
        email = clean_str(ws.cell(r, hidx["email"]).value)
        membership_no = clean_str(ws.cell(r, hidx["Membership No."]).value)
        member_since = ws.cell(r, hidx["Membership Date"]).value

        for owner_name in split_names(name_raw):
            person_id = get_or_create_person(owner_name, mobile1, mobile2, email)
            add_owner(
                site_id=site_id,
                person_id=person_id,
                membership_no=membership_no,
                member_since=member_since,
                is_current=not is_struck,
            )

    # Flag sites whose site_no was consolidated from A/B source variants.
    for site_id, originals in merged_from.items():
        site_row = site_id_to_row.get(site_id)
        if not site_row:
            continue
        orig_list = ", ".join(sorted(originals))
        comment = f"Site consolidated from source variants ({orig_list}) - verify co-owners and plot split"
        existing = site_row.get("FlagComment") or ""
        new_comment = (existing + " | " if existing else "") + comment
        site_row["FlaggedForAttention"] = "TRUE"
        site_row["FlagComment"] = new_comment[:500]
        site_row["FlaggedBy"] = MIGRATED_BY
        site_row["FlaggedAt"] = NOW

    # Ensure current owner exists for sites where struck owner was captured as historical.
    def get_person(person_id):
        for p in people:
            if p["PersonID"] == person_id:
                return p
        return None

    def current_owner_rows(site_id):
        return [o for o in owners_by_site.get(site_id, []) if o["IsCurrent"] == "TRUE"]

    mismatch_records = []

    for _, row in members_df.iterrows():
        m_site = normalize_site_no(row.get("Site", ""))
        m_phase = parse_phase(row.get("Phase", ""))
        if not m_site or m_phase not in (1, 2):
            continue

        key = f"{m_site}-P{m_phase}"
        site_id = site_key_to_id.get(key)
        if not site_id:
            continue

        member_name = normalize_name(row.get("Site Owner Name", ""))
        member_mobile = clean_mobile(row.get("Mobile", ""))
        member_mid = clean_str(row.get("Membership No.", ""))
        member_since = row.get("Member since", "")

        existing_current = current_owner_rows(site_id)

        # If no current owner exists, seed one from members sheet.
        if not existing_current and member_name:
            pid = get_or_create_person(member_name, member_mobile, "", "")
            add_owner(site_id, pid, member_mid, member_since, True)
            existing_current = current_owner_rows(site_id)

        if not existing_current:
            continue

        owner_names = set()
        owner_mobiles = set()
        for o in existing_current:
            person = get_person(o["PersonID"])
            if not person:
                continue
            owner_names.add(normalize_name(person.get("FullName", "")).lower())
            owner_mobiles.add(clean_mobile(person.get("Mobile1", "")))
            owner_mobiles.add(clean_mobile(person.get("Mobile2", "")))
        owner_mobiles.discard("")
        owner_mids = {clean_str(o.get("MembershipNo", "")) for o in existing_current}

        site_issues = []
        if member_name and not any(names_match(member_name, n) for n in owner_names):
            site_issues.append(f"Name mismatch with members sheet: '{member_name}'")
        if member_mobile and owner_mobiles and member_mobile not in owner_mobiles:
            site_issues.append(f"Mobile mismatch with members sheet: '{member_mobile}'")
        owner_mids_nonempty = {m for m in owner_mids if m}
        if member_mid and owner_mids_nonempty and member_mid not in owner_mids_nonempty:
            site_issues.append(f"MembershipNo mismatch with members sheet: '{member_mid}'")

        if site_issues:
            mismatch_records.append({
                "SiteID": site_id,
                "SiteNo": m_site,
                "Phase": m_phase,
                "Issues": " | ".join(site_issues),
            })
            for o in existing_current:
                o["FlaggedForAttention"] = "TRUE"
                o["FlagComment"] = " ; ".join(site_issues)[:500]
                o["FlaggedBy"] = MIGRATED_BY
                o["FlaggedAt"] = NOW

            site_row = site_id_to_row.get(site_id)
            if site_row:
                site_row["FlaggedForAttention"] = "TRUE"
                site_row["FlagComment"] = " | ".join(site_issues)[:500]
                site_row["FlaggedBy"] = MIGRATED_BY
                site_row["FlaggedAt"] = NOW

    # Cross-reference with old DB: flag attribute changes and sites absent from old DB.
    for site_row in sites:
        sno = site_row["SiteNo"]
        ph = site_row["Phase"]
        old = old_site_data.get((sno, ph))
        db_comments = []
        if old:
            if old["SiteType"] and old["SiteType"] != site_row.get("SiteType", ""):
                db_comments.append(
                    f"SiteType changed from old DB: {old['SiteType']} -> {site_row.get('SiteType', '')}"
                )
            try:
                sqft_old = float(old["Sizesqft"]) if str(old["Sizesqft"]).strip() not in ("", "nan") else None
                sqft_new = float(site_row.get("Sizesqft", "")) if str(site_row.get("Sizesqft", "")).strip() not in ("", "nan") else None
                if sqft_old is not None and sqft_new is not None and abs(sqft_old - sqft_new) > 1:
                    db_comments.append(
                        f"Sizesqft changed from old DB: {int(sqft_old)} -> {int(sqft_new)}"
                    )
            except Exception:
                pass
        else:
            db_comments.append("Site not present in previous database - verify site number")
        if db_comments:
            existing = site_row.get("FlagComment", "")
            new_comment = (existing + " | " if existing else "") + " | ".join(db_comments)
            site_row["FlaggedForAttention"] = "TRUE"
            site_row["FlagComment"] = new_comment[:500]
            if not site_row.get("FlaggedBy"):
                site_row["FlaggedBy"] = MIGRATED_BY
                site_row["FlaggedAt"] = NOW

    # Build Payments from bank receipts.
    payments = []

    site_to_current_owner = {}
    for o in owners:
        if o["IsCurrent"] == "TRUE" and o["SiteID"] not in site_to_current_owner:
            site_to_current_owner[o["SiteID"]] = o["OwnerID"]

    for _, row in bank_df.iterrows():
        tx_date = to_date(row.get("Txn Date", ""))
        description = clean_str(row.get("Description", ""))
        ref_no = clean_str(row.get("Ref No./Cheque No.", ""))
        branch = clean_str(row.get("Branch Code", ""))
        site_ph_raw = clean_str(row.get("Site/Ph", ""))
        bank_ref = f"Description: {description} | RefNo: {ref_no} | BranchCode: {branch}".strip()

        site_pairs = parse_site_phase_list(site_ph_raw)
        mapped_site_ids = []
        for site_no, phase in site_pairs:
            site_id = site_key_to_id.get(f"{site_no}-P{phase}")
            if site_id:
                mapped_site_ids.append(site_id)
            else:
                warnings.append(f"Bank receipt site not found: {site_no}/P{phase}")

        def add_payment(target_site_id, head_id, amount, flagged=False, flag_comment=""):
            if amount <= 0:
                return
            payments.append({
                "PaymentID": f"PAY{str(len(payments)+1).zfill(5)}",
                "SiteID": target_site_id,
                "OwnerID": site_to_current_owner.get(target_site_id, ""),
                "HeadID": head_id,
                "Amount": round(amount, 2),
                "Mode": "Bank transfer",
                "PaymentDate": tx_date,
                "ReceiptNo": ref_no,
                "BankRef": bank_ref,
                "ProofURL": "",
                "FlaggedForAttention": "TRUE" if flagged else "FALSE",
                "FlagComment": flag_comment[:500] if flag_comment else "",
                "FlaggedBy": MIGRATED_BY if flagged else "",
                "FlaggedAt": NOW if flagged else "",
                "RecordedBy": MIGRATED_BY,
                "RecordedAt": NOW,
                "IsDeleted": "FALSE",
                "CreatedBy": MIGRATED_BY,
                "CreatedAt": NOW,
                "ModifiedBy": "",
                "ModifiedAt": "",
            })

        amount_2526 = to_float(row.get("25-26", ""))
        amount_2627 = to_float(row.get("26-27", ""))
        amount_membership = to_float(row.get("Membership", ""))
        amount_additional = to_float(row.get("Additional", ""))
        paid_for = clean_str(row.get("Paid for", "")).lower()
        credit = to_float(row.get("Credit", ""))

        # Zombie payments: keep unmappable bank rows in Payments with blank site and a flag.
        if not mapped_site_ids:
            if site_pairs:
                zombie_reason = f"Zombie payment: site not found for parsed Site/Ph '{site_ph_raw}'. Update site mapping."
            else:
                zombie_reason = f"Zombie payment: invalid or blank Site/Ph '{site_ph_raw or 'blank'}'. Update site mapping."

            posted_any = False
            if amount_membership > 0:
                add_payment("", "H001", amount_membership, True, zombie_reason)
                posted_any = True
            if amount_2526 > 0:
                add_payment("", "H003", amount_2526, True, zombie_reason)
                posted_any = True
            if amount_2627 > 0:
                add_payment("", "H004", amount_2627, True, zombie_reason)
                posted_any = True

            if amount_membership == 0 and amount_2526 == 0 and amount_2627 == 0 and credit > 0:
                if "membership" in paid_for:
                    add_payment("", "H001", credit, True, zombie_reason)
                    posted_any = True
                elif "25-26" in paid_for and "26-27" not in paid_for:
                    add_payment("", "H003", credit, True, zombie_reason)
                    posted_any = True
                elif "26-27" in paid_for and "25-26" not in paid_for:
                    add_payment("", "H004", credit, True, zombie_reason)
                    posted_any = True
                elif "25-26" in paid_for and "26-27" in paid_for:
                    add_payment("", "H003", round(credit / 2.0, 2), True, zombie_reason)
                    add_payment("", "H004", round(credit / 2.0, 2), True, zombie_reason)
                    posted_any = True

            inferred_additional = 0.0
            if amount_additional > 0:
                inferred_additional = amount_additional
            elif "additional" in paid_for and credit > 0:
                inferred_additional = credit
            if inferred_additional > 0:
                add_payment("", additional_head_id, inferred_additional, True, zombie_reason)
                posted_any = True

            if not posted_any and credit > 0:
                add_payment("", additional_head_id, credit, True, zombie_reason + " Head inferred as Additional.")
            continue

        # Split explicit amount columns across mapped sites to avoid overcounting
        # when a single bank row references multiple Site/Ph values.
        n_sites = len(mapped_site_ids)
        per_site_membership = (amount_membership / n_sites) if n_sites and amount_membership > 0 else 0
        per_site_2526 = (amount_2526 / n_sites) if n_sites and amount_2526 > 0 else 0
        per_site_2627 = (amount_2627 / n_sites) if n_sites and amount_2627 > 0 else 0

        for target_site_id in mapped_site_ids:
            if per_site_membership > 0:
                add_payment(target_site_id, "H001", per_site_membership)
            if per_site_2526 > 0:
                add_payment(target_site_id, "H003", per_site_2526)
            if per_site_2627 > 0:
                add_payment(target_site_id, "H004", per_site_2627)

        # Fallback mapping when explicit columns are blank.
        if amount_membership == 0 and amount_2526 == 0 and amount_2627 == 0:
            per_site_credit = credit / len(mapped_site_ids) if mapped_site_ids else 0
            for target_site_id in mapped_site_ids:
                if "membership" in paid_for and per_site_credit > 0:
                    add_payment(target_site_id, "H001", per_site_credit)
                elif "25-26" in paid_for and "26-27" not in paid_for and per_site_credit > 0:
                    add_payment(target_site_id, "H003", per_site_credit)
                elif "26-27" in paid_for and "25-26" not in paid_for and per_site_credit > 0:
                    add_payment(target_site_id, "H004", per_site_credit)
                elif "25-26" in paid_for and "26-27" in paid_for and per_site_credit > 0:
                    add_payment(target_site_id, "H003", round(per_site_credit / 2.0, 2))
                    add_payment(target_site_id, "H004", round(per_site_credit / 2.0, 2))

        # Additional mapping: use explicit Additional column when present, otherwise infer from Paid for.
        inferred_additional = 0.0
        if amount_additional > 0:
            inferred_additional = amount_additional
        elif "additional" in paid_for and credit > 0:
            inferred_additional = credit

        if inferred_additional > 0:
            per_site_additional = inferred_additional / len(mapped_site_ids)
            for target_site_id in mapped_site_ids:
                add_payment(target_site_id, additional_head_id, per_site_additional)

    # Build CallLog from members list updates.
    call_logs = []

    for _, row in members_df.iterrows():
        m_site = normalize_site_no(row.get("Site", ""))
        m_phase = parse_phase(row.get("Phase", ""))
        if not m_site or m_phase not in (1, 2):
            continue

        site_id = site_key_to_id.get(f"{m_site}-P{m_phase}")
        if not site_id:
            continue

        summary_parts = []
        for label, col in [
            ("Call Notes", "Update - Call Notes by Priyanka"),
            ("MCOA Comments", "MCOA Comments for Priyanka"),
            ("Living in Bangalore", "Update - Living in Bangalore?"),
            ("Joining SGBM", "Update - Joining SGBM?"),
            ("Payment Reminder", "Update - Maintenance Payment Reminder for 2026-27"),
        ]:
            val = clean_str(row.get(col, ""))
            if val:
                summary_parts.append(f"{label}: {val}")

        if not summary_parts:
            continue

        owner_id = site_to_current_owner.get(site_id, "")
        call_logs.append({
            "LogID": f"L{str(len(call_logs)+1).zfill(4)}",
            "SiteID": site_id,
            "OwnerID": owner_id,
            "LogDate": TODAY,
            "CalledBy": "Priyanka",
            "Summary": " | ".join(summary_parts)[:2000],
            "FollowUpAction": "",
            "AssignedTo": "",
            "AssignedToName": "",
            "FollowUpDone": "FALSE",
            "DoneBy": "",
            "DoneAt": "",
            "LoggedAt": NOW,
            "IsDeleted": "FALSE",
            "CreatedBy": MIGRATED_BY,
            "CreatedAt": NOW,
            "ModifiedBy": "",
            "ModifiedAt": "",
        })

    agents_df = pd.DataFrame(columns=AGENTS_COLS)
    transfers_df = pd.DataFrame(columns=TRANSFERS_COLS)
    audit_df = pd.DataFrame(columns=AUDIT_COLS)

    sites_df = pd.DataFrame(sites, columns=SITES_COLS)
    people_df = pd.DataFrame(people, columns=PEOPLE_COLS)
    owners_df = pd.DataFrame(owners, columns=OWNERS_COLS)
    payments_df = pd.DataFrame(payments, columns=PAYMENTS_COLS)
    call_logs_df = pd.DataFrame(call_logs, columns=CALLLOG_COLS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sites_df.to_excel(writer, sheet_name="Sites", index=False)
        people_df.to_excel(writer, sheet_name="People", index=False)
        owners_df.to_excel(writer, sheet_name="Owners", index=False)
        agents_df.to_excel(writer, sheet_name="Agents", index=False)
        payments_df.to_excel(writer, sheet_name="Payments", index=False)
        payment_heads_df.to_excel(writer, sheet_name="PaymentHeads", index=False)
        transfers_df.to_excel(writer, sheet_name="Transfers", index=False)
        call_logs_df.to_excel(writer, sheet_name="CallLog", index=False)
        roles_df.to_excel(writer, sheet_name="Roles", index=False)
        audit_df.to_excel(writer, sheet_name="AuditLog", index=False)

    mismatch_df = pd.DataFrame(mismatch_records)
    mismatch_csv = output_path.with_name(output_path.stem + "_discrepancies.csv")
    mismatch_df.to_csv(mismatch_csv, index=False)

    warnings_txt = output_path.with_name(output_path.stem + "_warnings.txt")
    warnings_txt.write_text("\n".join(warnings) if warnings else "No warnings", encoding="utf-8")

    print("Fresh database written:", output_path)
    print("Discrepancy report:", mismatch_csv)
    print("Warnings report:", warnings_txt)
    print("Rows summary:")
    print("  Sites:", len(sites_df))
    print("  People:", len(people_df))
    print("  Owners:", len(owners_df))
    print("  Payments:", len(payments_df))
    print("  CallLog:", len(call_logs_df))
    print("  Discrepancies flagged:", len(mismatch_df))


def main():
    parser = argparse.ArgumentParser(description="Build fresh MyCiti Owners Database workbook")
    parser.add_argument(
        "--base-db",
        default=r"C:\Users\HP\Downloads\MyCiti Owners Database.xlsx",
        help="Current database workbook (Roles and PaymentHeads are retained)",
    )
    parser.add_argument(
        "--bank-csv",
        default=r"C:\Users\HP\Downloads\Bank Receipt from Dec 2024(Bank Recpts from 12Dec2024).csv",
        help="Bank receipts CSV",
    )
    parser.add_argument(
        "--owners-list",
        default=r"C:\Users\HP\Downloads\Myciti Owners List Phase 1 & Phase 2 .xlsx",
        help="Owners list workbook",
    )
    parser.add_argument(
        "--members-list",
        default=r"C:\Users\HP\Downloads\MyCiti Members List (Priyanka's Copy).xlsx",
        help="Members call list workbook",
    )
    parser.add_argument(
        "--output",
        default=r"C:\Users\HP\Downloads\MyCiti Owners Database - Fresh.xlsx",
        help="Output workbook path",
    )
    args = parser.parse_args()

    build_database(
        base_db_path=args.base_db,
        bank_csv_path=args.bank_csv,
        owners_list_path=args.owners_list,
        members_list_path=args.members_list,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
